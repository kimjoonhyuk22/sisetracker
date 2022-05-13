from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학
# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어" # B1 셀에 '국어' 입력

# 번호 영어 수학
# ws.move_range("C1:C11", rows=5, cols=-1)

# wb.save("sample_korean.xlsx")

ws.move_range("B1:C11",rows=0,cols=1)
ws["B1"].value="국어"

from random import*

index = 1
for x in range(2, 12): # 10 개 row
    for y in range(2, 3): # 10 개 column
        #ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample_korean1.xlsx")