# -*- coding: utf-8 -*-
from re import A
import time
import random
import requests
import json
import logging # 어떻게 하는지 모르겠지만 네이버 부동산 리퀴스트 모듈 안되는것 뚫음.
from openpyxl import load_workbook # 파일 불러오기

wb = load_workbook("sise.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

URLS=[] #리스트로 정의해주기
datass=[] #리스트로 정의해두기

# cell 데이터 불러오기 유알엘이 저장된값들 
for x in range(1, 517):  # 파일에 저장된값들 가져옴 1부터 시작해야됨. 셀에서는 
    for y in range(1, 2):
        sise = ws.cell(row=x, column=y).value  # 유알엘 링크를 가져옴
        URLS.append(sise) # 유알엘 링크를 마늘어둔 유알엘스 리스트에 어팬더로 하나씩 저장함. 나중에 유알엘스를 이용해서 값을 가져올것임.
    

# print(URLS,end=" ")

# URL = ["https://new.land.naver.com/api/complexes/single-markers/2.0?cortarNo=1147010200&zoom=18&priceType=RETAIL&markerId=659&markerType=COMPLEX&selectedComplexNo&selectedComplexBuildingNo&fakeComplexMarker&realEstateType=APT%3AABYG%3AJGC&tradeType=&tag=%3A%3A%3A%3A%3A%3A%3A%3A&rentPriceMin=0&rentPriceMax=900000000&priceMin=0&priceMax=900000000&areaMin=95.9&areaMax=125.6&oldBuildYears&recentlyBuildYears&minHouseHoldCount=300&maxHouseHoldCount&showArticle=false&sameAddressGroup=true&minMaintenanceCost&maxMaintenanceCost&directions=&leftLon=126.8637504&rightLon=126.8706168&topLat=37.5293148&bottomLat=37.5265495","https://new.land.naver.com/api/complexes/single-markers/2.0?cortarNo=1165010700&zoom=19&priceType=RETAIL&markerId&markerType&selectedComplexNo&selectedComplexBuildingNo&fakeComplexMarker&realEstateType=APT%3AABYG%3AJGC&tradeType=&tag=%3A%3A%3A%3A%3A%3A%3A%3A&rentPriceMin=0&rentPriceMax=900000000&priceMin=0&priceMax=900000000&areaMin=95.9&areaMax=125.6&oldBuildYears&recentlyBuildYears&minHouseHoldCount=300&maxHouseHoldCount&showArticle=false&sameAddressGroup=true&minMaintenanceCost&maxMaintenanceCost&directions=&leftLon=126.9917322&rightLon=126.9951654&topLat=37.5067621&bottomLat=37.5053791"]
# 아래 있는 파람과 해더는 왜쓰는지 모르겠으나 중요함.
param = {
    'hscpNo': '19672',
    'tradTpCd': 'A1',
    'order': 'date_',
    'showR0': 'N',
}

header = {
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.220 Whale/1.3.51.7 Safari/537.36",
    'Referer': 'https://m.land.naver.com/'
}

logging.basicConfig(level=logging.INFO)
page = 0


for a in range(0,516): #유알엘스에 516개의 링크가 저장되어 있고 하나씩 하나씩 꺼내서 존슨형태로 저장해서 data에 리스트형태로 저장함.
    t=random.random()
    time.sleep(t)
    print(a,t)
    resp = requests.get(URLS[a], params=param, headers=header)
    data = json.loads(resp.text) #이때 저장되는 값들은 [{키:벨류}]의 형태로 리스트 안에 딕셔너리 값으로 들어가있음.

    datas=data[0].get('complexName') #위에 저장된 data는 []리스트 형태로 되어 있으니[0]번째에 모든값이 딕셔너리에 저장되어 있고 딕셔너리에서 키값으로 찾으면 값이 없는게 생김
    datass.append(datas) #값이 없는게 생기면 에러가 발생하기 때문에 get(키값)함수를 이용하여 데이터를 값을 가지고 오고 이값을 데이터스스에 저장시킴.
    datas=data[0].get('minDealPrice')
    datass.append(datas)
    datas=data[0].get('maxDealPrice')
    datass.append(datas)
    datas=data[0].get('minLeasePrice')
    datass.append(datas)
    datas=data[0].get('maxLeasePrice')
    datass.append(datas)
    datas=data[0].get('medianDealPrice')
    datass.append(datas)

# print(datass[::6])
# print(datass[1::6])
# print(datass[2::6])
# print(datass[3::6])
# print(datass[4::6])
# print(datass[5::6]) #넘 길어서 출력안되나 싶음.


# 1줄씩 데이터 넣기
ws.append(["이름", "최소매매", "최대매매", "최소전세", "최대전세", "중간값"]) # 엑셀에 목차를 추가함.
for i in range(0,516): # 데이터를 엑셀에 넣어줄것인데 6개씩 묶어서 저장하기 위해서 6*i를 이용함. 그럼 한줄씩 띄어서 출력이됨.
    ws.append([datass[6*i],datass[6*i+1],datass[6*i+2],datass[6*i+3],datass[6*i+4],datass[6*i+5]])
wb.save("sesitracker.xlsx") # 마지막으로 저장하면끝.


##########################느낀점 : 모듈안의 함수에 대한 정의를 알수가 없다. 어떤 파라미터를 적어야하고 적으면 어떤 타입으로 출력되는지를 알아야지 잘쓸것 같은데. 그리고 함수별 추가 함수쓸수 있는것은 무엇인지 어팬던드가 되는지 안되는지 같은것이 명확하지가 않음.
###하지만 모를때 마다 검색하며 오류 검색하며 하니 됨....이러다보니 자신은 없지만 그냥 노가다로 됨.

    # if data[0]["medianDealPrice"] != None: # 의미 없는 데이터는 skip
    #     print(data[0]["medianDealPrice"])
    # else:
    #     data[0]["medianDealPrice"] = 0

    # print(data[0]["complexName"],data[0]["maxDealPrice"],data[0]["maxDealPrice"],data[0]["minLeasePrice"],data[0]["maxLeasePrice"],data[0]["medianDealPrice"])
        # continue
 # medianDealPrice 이 없는게 있어서 에러남.조건문써야됨.


# while True:
#     page += 1
#     param['page'] = page

#     resp = requests.get(URL, params=param, headers=header)
#     if resp.status_code != 200:
#         logging.error('invalid status: %d' % resp.status_code)
#         break

#     data = json.loads(resp.text)
#     result = data['result']
#     if result is None:
#         logging.error('no result')
#         break
    
#     for item in result['list']:
#         if float(item['spc2']) < 80 or float(item['spc2']) > 85:
#             continue
#         logging.info('[%s] %s %s층 %s만원' % (item['tradTpNm'], item['bildNm'], item['flrInfo'], item['prcInfo']))
    
#     if result['moreDataYn'] == 'N':
#         break

