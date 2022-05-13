# -*- coding: utf-8 -*-

import requests
import json
import logging
from openpyxl import load_workbook # 파일 불러오기

wb = load_workbook("sise.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

URLS=[]
datass=[]

# cell 데이터 불러오기
for x in range(1, 516):
    for y in range(1, 2):
        sise = ws.cell(row=x, column=y).value
        URLS.append(sise)
    

# print(URLS,end=" ")

# URL = ["https://new.land.naver.com/api/complexes/single-markers/2.0?cortarNo=1147010200&zoom=18&priceType=RETAIL&markerId=659&markerType=COMPLEX&selectedComplexNo&selectedComplexBuildingNo&fakeComplexMarker&realEstateType=APT%3AABYG%3AJGC&tradeType=&tag=%3A%3A%3A%3A%3A%3A%3A%3A&rentPriceMin=0&rentPriceMax=900000000&priceMin=0&priceMax=900000000&areaMin=95.9&areaMax=125.6&oldBuildYears&recentlyBuildYears&minHouseHoldCount=300&maxHouseHoldCount&showArticle=false&sameAddressGroup=true&minMaintenanceCost&maxMaintenanceCost&directions=&leftLon=126.8637504&rightLon=126.8706168&topLat=37.5293148&bottomLat=37.5265495","https://new.land.naver.com/api/complexes/single-markers/2.0?cortarNo=1165010700&zoom=19&priceType=RETAIL&markerId&markerType&selectedComplexNo&selectedComplexBuildingNo&fakeComplexMarker&realEstateType=APT%3AABYG%3AJGC&tradeType=&tag=%3A%3A%3A%3A%3A%3A%3A%3A&rentPriceMin=0&rentPriceMax=900000000&priceMin=0&priceMax=900000000&areaMin=95.9&areaMax=125.6&oldBuildYears&recentlyBuildYears&minHouseHoldCount=300&maxHouseHoldCount&showArticle=false&sameAddressGroup=true&minMaintenanceCost&maxMaintenanceCost&directions=&leftLon=126.9917322&rightLon=126.9951654&topLat=37.5067621&bottomLat=37.5053791"]

param = {
    'hscpNo': '19672',
    'tradTpCd': 'A1',
    'order': 'date_',
    'showR0': 'N',
}

header = {
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.220 Whale/1.3.51.7 Safari/537.36",
    'Referer': 'https://m.land.naver.com/'
}

logging.basicConfig(level=logging.INFO)
page = 0

for a in range(0,9):
    resp = requests.get(URLS[a], params=param, headers=header)
    data = json.loads(resp.text)

    datas=data[0].get('complexName') 
    datass.append(datas)
    datas=data[0].get('minDealPrice')
    datass.append(datas)
    datas=data[0].get('maxDealPrice')
    datass.append(datas)
    datas=data[0].get('minLeasePrice')
    datass.append(datas)
    datas=data[0].get('maxLeasePrice')
    datass.append(datas)
    datas=data[0].get('medianDealPrice')
    datass.append(datas)

print(datass[::6])
print(datass[1::6])
print(datass[2::6])
print(datass[3::6])
print(datass[4::6])
print(datass[5::6])


    # if data[0]["medianDealPrice"] != None: # 의미 없는 데이터는 skip
    #     print(data[0]["medianDealPrice"])
    # else:
    #     data[0]["medianDealPrice"] = 0

    # print(data[0]["complexName"],data[0]["maxDealPrice"],data[0]["maxDealPrice"],data[0]["minLeasePrice"],data[0]["maxLeasePrice"],data[0]["medianDealPrice"])
        # continue
 # medianDealPrice 이 없는게 있어서 에러남.조건문써야됨. 겟쓰면 없으면 없는데로 있느면 있는데로 다른거 가지고오는것 쓰면 없으면 안가져옴


# while True:
#     page += 1
#     param['page'] = page

#     resp = requests.get(URL, params=param, headers=header)
#     if resp.status_code != 200:
#         logging.error('invalid status: %d' % resp.status_code)
#         break

#     data = json.loads(resp.text)
#     result = data['result']
#     if result is None:
#         logging.error('no result')
#         break
    
#     for item in result['list']:
#         if float(item['spc2']) < 80 or float(item['spc2']) > 85:
#             continue
#         logging.info('[%s] %s %s층 %s만원' % (item['tradTpNm'], item['bildNm'], item['flrInfo'], item['prcInfo']))
    
#     if result['moreDataYn'] == 'N':
#         break

